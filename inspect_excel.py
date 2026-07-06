import openpyxl
import json
from collections import defaultdict

wb = openpyxl.load_workbook(r'c:\laragon\www\solca-parte-diario\DR. AUCAPIÑA.xlsx', data_only=True)
output = {}

for sheet_name in wb.sheetnames[:5]:
    sheet = wb[sheet_name]
    output[sheet_name] = {
        'merged_cells': [str(mc) for mc in sheet.merged_cells.ranges],
        'rows': []
    }
    for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=55):
        row_data = []
        for cell in row:
            if cell.value is not None:
                row_data.append(f"{cell.coordinate}: {cell.value}")
        if row_data:
            output[sheet_name]['rows'].append(row_data)

with open(r'c:\laragon\www\solca-parte-diario\excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)
