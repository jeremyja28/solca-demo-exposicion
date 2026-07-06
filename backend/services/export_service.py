import io
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generar_pdf_parte_diario(registros, medico_nombre, fecha):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    # Encabezado
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 50, "SOCIEDAD DE LUCHA CONTRA EL CÁNCER — SOLCA")
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 70, "Parte Diario Médico")
    c.drawString(50, height - 90, f"Médico: {medico_nombre}   |   Fecha: {fecha}")
    
    c.setStrokeColorRGB(0, 0.2, 0.4) # #003366
    c.line(50, height - 100, width - 50, height - 100)

    # Tabla Header
    y = height - 130
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0, 0.2, 0.4)
    c.rect(50, y, width - 100, 20, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    
    headers = [
        ("N°HC", 55), ("Nombres", 100), ("Edad", 250), ("Sexo", 280),
        ("Diagnóstico", 320), ("CIE10", 450), ("Convenio", 500),
        ("Especialidad", 570), ("Act", 640), ("Tipo", 680),
        ("QT", 720), ("QX", 740), ("Q", 760), ("E", 780)
    ]
    for text, x in headers:
        c.drawString(x, y + 6, text)

    # Filas
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0, 0, 0)
    y -= 20
    
    for i, reg in enumerate(registros):
        if y < 50:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 8)
            
        if i % 2 != 0:
            c.setFillColorRGB(0.96, 0.97, 0.98) # #F5F7FA
            c.rect(50, y, width - 100, 20, fill=1, stroke=0)
            
        c.setFillColorRGB(0, 0, 0)
        c.drawString(55, y + 6, str(reg.N_HC))
        nombres = f"{reg.Apellidos} {reg.Nombres}"[:25]
        c.drawString(100, y + 6, nombres)
        c.drawString(250, y + 6, str(reg.Edad))
        c.drawString(280, y + 6, str(reg.Sexo))
        c.drawString(320, y + 6, str(reg.Diagnostico)[:25])
        c.drawString(450, y + 6, str(reg.CIE10))
        c.drawString(500, y + 6, str(reg.Convenio)[:12])
        
        if reg.Complemento:
            c.drawString(570, y + 6, str(reg.Complemento.EspecialidadId))
            c.drawString(640, y + 6, str(reg.Complemento.ActividadId))
            c.drawString(680, y + 6, reg.Complemento.TipoConsulta[:4])
            c.drawString(720, y + 6, "SÍ" if reg.Complemento.Pre_QT else "")
            c.drawString(740, y + 6, "SÍ" if reg.Complemento.Pre_QX else "")
            c.drawString(760, y + 6, "SÍ" if reg.Complemento.Quimio else "")
            c.drawString(780, y + 6, "SÍ" if reg.Complemento.EKG else "")
            
        y -= 20

    c.save()
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_parte_diario(registros, medico_nombre, fecha):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parte Diario"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_font = Font(bold=True, size=14, color="003366")
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    ws.merge_cells("A1:N1")
    cell = ws["A1"]
    cell.value = "SOLCA — Parte Diario Médico"
    cell.font = title_font

    ws["A2"] = f"Médico: {medico_nombre}  |  Fecha: {fecha}"
    
    headers = ["N° HC", "Nombres Completos", "Edad", "Sexo", "Diagnóstico", "CIE10", "Convenio",
               "Especialidad", "Actividad", "Tipo Consulta", "PRE QT", "PRE QX", "QUIMIO", "EKG"]
    
    ws.append([]) # separador
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col)
        c.font = header_font
        c.fill = header_fill

    tot_qt, tot_qx, tot_q, tot_e = 0, 0, 0, 0
    
    for row_idx, reg in enumerate(registros, start=5):
        qt = "SÍ" if reg.Complemento and reg.Complemento.Pre_QT else ""
        qx = "SÍ" if reg.Complemento and reg.Complemento.Pre_QX else ""
        qu = "SÍ" if reg.Complemento and reg.Complemento.Quimio else ""
        ek = "SÍ" if reg.Complemento and reg.Complemento.EKG else ""
        
        if qt: tot_qt += 1
        if qx: tot_qx += 1
        if qu: tot_q += 1
        if ek: tot_e += 1

        row_data = [
            reg.N_HC,
            f"{reg.Apellidos} {reg.Nombres}",
            reg.Edad,
            reg.Sexo,
            reg.Diagnostico,
            reg.CIE10,
            reg.Convenio,
            reg.Complemento.EspecialidadId if reg.Complemento else "",
            reg.Complemento.ActividadId if reg.Complemento else "",
            reg.Complemento.TipoConsulta if reg.Complemento else "",
            qt, qx, qu, ek
        ]
        ws.append(row_data)
        
        if (row_idx - 4) % 2 != 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    ws.append(["TOTALES", "", "", "", "", "", "", "", "", "", tot_qt, tot_qx, tot_q, tot_e])

    # Autoajustar
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_concentrado_mensual(registros, fecha_inicio, fecha_fin):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Detalle"
    headers1 = ["MédicoId", "EspecialidadId", "ActividadId", "N°HC", "Apellidos Nombres", "Edad", "Sexo", 
                "Diagnóstico", "Convenio", "Tipo Consulta", "PRE QT", "PRE QX", "QUIMIO", "EKG"]
    ws1.append(headers1)
    
    # Simple hoja 1
    for r in registros:
        ws1.append([
            r.MedicoId,
            r.Complemento.EspecialidadId if r.Complemento else "",
            r.Complemento.ActividadId if r.Complemento else "",
            r.N_HC,
            f"{r.Apellidos} {r.Nombres}",
            r.Edad, r.Sexo, r.Diagnostico, r.Convenio,
            r.Complemento.TipoConsulta if r.Complemento else "",
            "SÍ" if r.Complemento and r.Complemento.Pre_QT else "",
            "SÍ" if r.Complemento and r.Complemento.Pre_QX else "",
            "SÍ" if r.Complemento and r.Complemento.Quimio else "",
            "SÍ" if r.Complemento and r.Complemento.EKG else "",
        ])

    # Hoja 2
    ws2 = wb.create_sheet("Resumen por Médico")
    ws2.append(["Médico", "Total Pacientes", "Primera Vez", "Subsecuente", "PRE QT", "PRE QX", "QUIMIO", "EKG"])
    
    # Agrupación simple
    medicos = {}
    for r in registros:
        mid = r.MedicoId
        if mid not in medicos:
            medicos[mid] = {"tot": 0, "pv": 0, "sub": 0, "qt": 0, "qx": 0, "qu": 0, "ek": 0}
        
        m = medicos[mid]
        m["tot"] += 1
        if r.Complemento:
            if r.Complemento.TipoConsulta == "PRIMERA_VEZ": m["pv"] += 1
            if r.Complemento.TipoConsulta == "SUBSECUENTE": m["sub"] += 1
            if r.Complemento.Pre_QT: m["qt"] += 1
            if r.Complemento.Pre_QX: m["qx"] += 1
            if r.Complemento.Quimio: m["qu"] += 1
            if r.Complemento.EKG: m["ek"] += 1

    for mid, m in medicos.items():
        ws2.append([mid, m["tot"], m["pv"], m["sub"], m["qt"], m["qx"], m["qu"], m["ek"]])

    # Hoja 3
    ws3 = wb.create_sheet("Resumen por Especialidad")
    ws3.append(["EspecialidadId", "Total", "Primera Vez", "Subsecuente", "PRE QT", "PRE QX", "QUIMIO", "EKG"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# --- LÓGICA DE EXCEL MENSUAL ---

import collections

def _get_age_group_index(age: int) -> int:
    if age <= 4: return 0
    if age <= 9: return 1
    if age <= 14: return 2
    if age <= 19: return 3
    if age <= 35: return 4
    if age <= 49: return 5
    if age <= 64: return 6
    return 7

_AGE_GROUPS = ["1-4", "5-9", "10-14", "15-19", "20-35", "36-49", "50-64", "65 Y +"]

def _aplicar_bordes_y_alineacion(ws, min_row, max_row, min_col, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

def _crear_hoja_plan_diario(ws, registros_tuplas, medico_nombre, anio, mes):
    ws.title = "Plan Diario"
    
    # Encabezados
    ws.merge_cells('A1:S1')
    ws['A1'] = f"PLANILLA DE CONCENTRACIÓN MENSUAL CONSULTA EXTERNA - {mes}/{anio}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"MÉDICO: {medico_nombre}"
    ws['A3'].font = Font(bold=True)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells('A5:A7')
    ws['A5'] = "DIA LABORADO (FECHA)"
    
    ws.merge_cells('B5:S5')
    ws['B5'] = "MORBILIDAD"
    
    ws.merge_cells('B6:C6')
    ws['B6'] = "SEXO"
    ws['B7'] = "H"
    ws['C7'] = "M"

    ws.merge_cells('D6:K6')
    ws['D6'] = "PRIMERAS CONSULTAS"
    for i, ag in enumerate(_AGE_GROUPS):
        ws.cell(row=7, column=4+i, value=ag)

    ws.merge_cells('L6:S6')
    ws['L6'] = "SUBSECUENTES"
    for i, ag in enumerate(_AGE_GROUPS):
        ws.cell(row=7, column=12+i, value=ag)

    for r in range(5, 8):
        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Agrupación de datos
    datos_por_dia = collections.defaultdict(lambda: {
        'H': 0, 'M': 0,
        'P': [0]*8, 'S': [0]*8
    })

    for reg, fecha in registros_tuplas:
        if not fecha: continue
        dia = fecha if isinstance(fecha, str) else fecha.isoformat()
        
        sexo = reg.Sexo.upper() if reg.Sexo else ""
        is_m = sexo.startswith('M') and not sexo.startswith('MUJ')
        if sexo == 'MASCULINO' or sexo == 'H' or sexo == 'M' and is_m:
            sex_key = 'H'
        else:
            sex_key = 'M' # Mujer
            
        if sexo == 'M': sex_key = 'H'
        if sexo == 'F': sex_key = 'M'
        
        datos_por_dia[dia][sex_key] += 1

        tipo = "SUBSECUENTE"
        if reg.Complemento and reg.Complemento.TipoConsulta:
            tipo = reg.Complemento.TipoConsulta.upper()
            
        age_idx = _get_age_group_index(reg.Edad)
        
        if "PRIMER" in tipo or "NUEV" in tipo:
            datos_por_dia[dia]['P'][age_idx] += 1
        else:
            datos_por_dia[dia]['S'][age_idx] += 1

    dias = sorted(list(datos_por_dia.keys()))
    current_row = 8
    for dia in dias:
        data = datos_por_dia[dia]
        ws.cell(row=current_row, column=1, value=dia)
        ws.cell(row=current_row, column=2, value=data['H'])
        ws.cell(row=current_row, column=3, value=data['M'])
        
        for i in range(8):
            ws.cell(row=current_row, column=4+i, value=data['P'][i])
            ws.cell(row=current_row, column=12+i, value=data['S'][i])
        current_row += 1

    _aplicar_bordes_y_alineacion(ws, 8, current_row-1, 1, 19)
    ws.column_dimensions['A'].width = 15

def _crear_hoja_concentrado(ws, registros_tuplas, medico_nombre):
    ws.title = "Concentrado Especialidad"
    
    ws.merge_cells('A1:AG1')
    ws['A1'] = "CONCENTRADO POR ESPECIALIDAD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells('A3:A5')
    ws['A3'] = "ESPECIALIDAD"
    
    ws.merge_cells('B3:Q3')
    ws['B3'] = "PRIMERAS CONSULTAS"
    ws.merge_cells('B4:I4')
    ws['B4'] = "HOMBRES"
    ws.merge_cells('J4:Q4')
    ws['J4'] = "MUJERES"

    ws.merge_cells('R3:AG3')
    ws['R3'] = "SUBSECUENTES"
    ws.merge_cells('R4:Y4')
    ws['R4'] = "HOMBRES"
    ws.merge_cells('Z4:AG4')
    ws['Z4'] = "MUJERES"

    for base_col in [2, 10, 18, 26]:
        for i, ag in enumerate(_AGE_GROUPS):
            ws.cell(row=5, column=base_col+i, value=ag)

    for r in range(3, 6):
        for c in range(1, 34):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    datos_por_esp = collections.defaultdict(lambda: {
        'P_H': [0]*8, 'P_M': [0]*8,
        'S_H': [0]*8, 'S_M': [0]*8
    })

    for reg, fecha in registros_tuplas:
        esp = reg.Complemento.EspecialidadId if reg.Complemento and reg.Complemento.EspecialidadId else "Sin Especialidad"
        
        sexo = reg.Sexo.upper() if reg.Sexo else ""
        if sexo == 'M' or sexo == 'H' or sexo == 'MASCULINO': sex_key = 'H'
        else: sex_key = 'M'
        
        tipo = "SUBSECUENTE"
        if reg.Complemento and reg.Complemento.TipoConsulta:
            tipo = reg.Complemento.TipoConsulta.upper()
            
        age_idx = _get_age_group_index(reg.Edad)
        
        prefix = 'P' if "PRIMER" in tipo or "NUEV" in tipo else 'S'
        datos_por_esp[esp][f"{prefix}_{sex_key}"][age_idx] += 1

    current_row = 6
    for esp, data in datos_por_esp.items():
        ws.cell(row=current_row, column=1, value=esp)
        for i in range(8):
            ws.cell(row=current_row, column=2+i, value=data['P_H'][i])
            ws.cell(row=current_row, column=10+i, value=data['P_M'][i])
            ws.cell(row=current_row, column=18+i, value=data['S_H'][i])
            ws.cell(row=current_row, column=26+i, value=data['S_M'][i])
        current_row += 1

    _aplicar_bordes_y_alineacion(ws, 6, current_row-1, 1, 33)
    ws.column_dimensions['A'].width = 30

def _crear_hoja_diagnosticos(ws, registros_tuplas):
    ws.title = "Diagnósticos CIE-10"
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "CAUSAS DE MORBILIDAD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ["N°", "CIE-10", "CAUSA DE MORBILIDAD", "TOTAL CONSULTAS", "HOMBRES", "MUJERES"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    datos_diag = collections.defaultdict(lambda: {'desc': '', 'H': 0, 'M': 0})

    for reg, fecha in registros_tuplas:
        if not reg.CIE10: continue
        
        codigos = [c.strip() for c in reg.CIE10.split(';')]
        descripciones = [d.strip() for d in reg.Diagnostico.split(';')] if reg.Diagnostico else []
        
        sexo = reg.Sexo.upper() if reg.Sexo else ""
        if sexo == 'M' or sexo == 'H' or sexo == 'MASCULINO': sex_key = 'H'
        else: sex_key = 'M'

        for idx, codigo in enumerate(codigos):
            if not codigo: continue
            desc = descripciones[idx] if idx < len(descripciones) else "NO ESPECIFICADO"
            
            if not datos_diag[codigo]['desc']:
                datos_diag[codigo]['desc'] = desc
            
            datos_diag[codigo][sex_key] += 1

    current_row = 4
    for i, (codigo, data) in enumerate(datos_diag.items(), start=1):
        ws.cell(row=current_row, column=1, value=i)
        ws.cell(row=current_row, column=2, value=codigo)
        ws.cell(row=current_row, column=3, value=data['desc'])
        ws.cell(row=current_row, column=4, value=data['H'] + data['M'])
        ws.cell(row=current_row, column=5, value=data['H'])
        ws.cell(row=current_row, column=6, value=data['M'])
        current_row += 1

    _aplicar_bordes_y_alineacion(ws, 4, current_row-1, 1, 6)
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 70


def generar_excel_concentrado_mensual(registros_tuplas, medico_nombre, anio, mes):
    wb = Workbook()
    
    # 1. Hoja Plan Diario
    ws1 = wb.active
    _crear_hoja_plan_diario(ws1, registros_tuplas, medico_nombre, anio, mes)
    
    # 2. Hoja Concentrado Especialidad
    ws2 = wb.create_sheet()
    _crear_hoja_concentrado(ws2, registros_tuplas, medico_nombre)
    
    # 3. Hoja Diagnosticos
    ws3 = wb.create_sheet()
    _crear_hoja_diagnosticos(ws3, registros_tuplas)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

